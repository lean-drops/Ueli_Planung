import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

def generate_excel(employee_name):
    wb = Workbook()
    ws = wb.active

    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30

    # Überschriften
    ws['A1'] = 'Datum'
    ws['B1'] = 'Verfügbarkeit'

    # Datenvalidierung für Verfügbarkeit
    dv = DataValidation(type="list", formula1='"Verfügbar,Nicht verfügbar"', showDropDown=True)
    ws.add_data_validation(dv)
    for row in range(2, 32):  # Für 30 Tage als Beispiel
        ws[f'B{row}'] = 'Bitte wählen'
        dv.add(ws[f'B{row}'])

    # Datei speichern
    filename = f"{employee_name}_Verfügbarkeiten.xlsx"
    wb.save(filename)

def add_employee():
    employee_name = simpledialog.askstring("Mitarbeiter hinzufügen", "Name des Mitarbeiters:")
    if employee_name:
        employees.append(employee_name)
        update_employee_list()

def generate_excel_for_all():
    for employee in employees:
        generate_excel(employee)
    messagebox.showinfo("Erfolg", "Excel-Dateien wurden erfolgreich generiert.")

def update_employee_list():
    employee_list_var.set('\n'.join(employees))

# GUI-Initialisierung
root = tk.Tk()
root.title("Excel-Generator für Mitarbeiter")

# Mitarbeiterliste
employees = []
employee_list_var = tk.StringVar()
employee_label = tk.Label(root, textvariable=employee_list_var, justify=tk.LEFT)
employee_label.pack()

# Buttons
add_button = tk.Button(root, text="Mitarbeiter hinzufügen", command=add_employee)
add_button.pack()

generate_button = tk.Button(root, text="Excel-Dateien generieren", command=generate_excel_for_all)
generate_button.pack()

root.mainloop()
